import csv
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def convert_txt_to_xlsx(input_txt_path, output_xlsx_path):
    if not os.path.exists(input_txt_path):
        raise FileNotFoundError(f"Input file not found: {input_txt_path}")

    with open(input_txt_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError("The input file is empty.")

    header = rows[0]
    data_rows = rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write header
    for col_idx, col_name in enumerate(header, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cleaned_value = value.strip()

            if cleaned_value == "":
                ws.cell(row=row_idx, column=col_idx, value="")
                continue

            try:
                if "." in cleaned_value:
                    ws.cell(row=row_idx, column=col_idx, value=float(cleaned_value))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=int(cleaned_value))
            except ValueError:
                ws.cell(row=row_idx, column=col_idx, value=cleaned_value)

    # Styles
    black_fill = PatternFill(fill_type="solid", fgColor="000000")
    white_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # Header style
    for cell in ws[1]:
        cell.fill = black_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Green columns based on your example
    green_columns_by_name = {"OD_IN", "AREA_SQIN", "R25", "R75"}

    green_col_indices = []
    for idx, name in enumerate(header, start=1):
        if str(name).strip() in green_columns_by_name:
            green_col_indices.append(idx)

    # Data styling
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in green_col_indices:
                cell.fill = green_fill

    # Filter and freeze
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # Column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_xlsx_path)


class TxtToXlsxGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("TXT to XLSX Converter")
        self.root.geometry("700x220")
        self.root.resizable(False, False)

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self.build_gui()

    def build_gui(self):
        title_label = tk.Label(
            self.root,
            text="Convert TXT Conductor Data to Excel",
            font=("Arial", 14, "bold")
        )
        title_label.pack(pady=15)

        # Input row
        input_frame = tk.Frame(self.root)
        input_frame.pack(fill="x", padx=20, pady=8)

        tk.Label(input_frame, text="Input TXT File:", width=15, anchor="w").pack(side="left")
        tk.Entry(input_frame, textvariable=self.input_path, width=60).pack(side="left", padx=5)
        tk.Button(input_frame, text="Browse", width=10, command=self.browse_input).pack(side="left")

        # Output row
        output_frame = tk.Frame(self.root)
        output_frame.pack(fill="x", padx=20, pady=8)

        tk.Label(output_frame, text="Save XLSX As:", width=15, anchor="w").pack(side="left")
        tk.Entry(output_frame, textvariable=self.output_path, width=60).pack(side="left", padx=5)
        tk.Button(output_frame, text="Browse", width=10, command=self.browse_output).pack(side="left")

        # Convert button
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=20)

        tk.Button(
            button_frame,
            text="Convert",
            width=20,
            height=2,
            command=self.convert_file
        ).pack()

    def browse_input(self):
        file_path = filedialog.askopenfilename(
            title="Select TXT File",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_path.set(file_path)

            # Auto-suggest output path
            base_name = os.path.splitext(file_path)[0]
            suggested_output = base_name + ".xlsx"
            if not self.output_path.get():
                self.output_path.set(suggested_output)

    def browse_output(self):
        file_path = filedialog.asksaveasfilename(
            title="Save XLSX File As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            self.output_path.set(file_path)

    def convert_file(self):
        input_file = self.input_path.get().strip()
        output_file = self.output_path.get().strip()

        if not input_file:
            messagebox.showerror("Error", "Please select an input TXT file.")
            return

        if not output_file:
            messagebox.showerror("Error", "Please select where to save the XLSX file.")
            return

        try:
            convert_txt_to_xlsx(input_file, output_file)
            messagebox.showinfo("Success", f"Excel file created successfully:\n{output_file}")
        except Exception as e:
            messagebox.showerror("Conversion Failed", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = TxtToXlsxGUI(root)
    root.mainloop()