import os
import re
import subprocess
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SLOTS = OrderedDict(
    [
        ("P1_2026", ("P1", "2026")),
        ("P1_2030", ("P1", "2030")),
        ("P2_2026", ("P2", "2026")),
        ("P2_2030", ("P2", "2030")),
    ]
)

CONTINGENCY_RE = re.compile(r"^\s*CONTINGENCY\s+['\"](?P<name>[^'\"]+)['\"]\s*$", re.IGNORECASE)
END_RE = re.compile(r"^\s*END\s*$", re.IGNORECASE)
WHITESPACE_RE = re.compile(r"\s+")


def normalize_action_line(line):
    return WHITESPACE_RE.sub(" ", line.strip()).upper()


def join_actions(actions):
    return "\n".join(actions)


def format_timestamp(timestamp):
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def find_con_files(folder):
    folder_path = Path(folder)
    slot_matches = {slot: [] for slot in SLOTS}

    for entry in folder_path.iterdir():
        if not entry.is_file() or entry.suffix.lower() != ".con":
            continue

        upper_name = entry.name.upper()
        for slot, (group, year) in SLOTS.items():
            if group in upper_name and year in upper_name:
                slot_matches[slot].append(entry)

    selected = {}
    warnings = []
    missing = []

    for slot, matches in slot_matches.items():
        if not matches:
            missing.append(slot)
            continue

        matches.sort(key=lambda path: path.stat().st_mtime, reverse=True)
        selected[slot] = str(matches[0])

        if len(matches) > 1:
            chosen = matches[0]
            warnings.append(
                f"{slot}: multiple matches found; using newest file "
                f"{chosen.name} ({format_timestamp(chosen.stat().st_mtime)})"
            )

    return {
        "selected": selected,
        "warnings": warnings,
        "missing": missing,
        "all_matches": {slot: [str(path) for path in paths] for slot, paths in slot_matches.items()},
    }


def parse_con_file(path):
    contingencies = OrderedDict()
    current_name = None
    current_actions = []

    with open(path, "r", encoding="utf-8-sig", errors="replace") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\r\n")
            match = CONTINGENCY_RE.match(line)
            if match:
                if current_name is not None:
                    contingencies[current_name] = current_actions
                current_name = match.group("name")
                current_actions = []
                continue

            if current_name is None:
                continue

            if END_RE.match(line):
                contingencies[current_name] = current_actions
                current_name = None
                current_actions = []
                continue

            current_actions.append(line)

    if current_name is not None:
        contingencies[current_name] = current_actions

    return contingencies


def compare_sets(dict_2026, dict_2030):
    names_2026 = set(dict_2026)
    names_2030 = set(dict_2030)

    unchanged = []
    modified = []
    removed = []
    added = []

    for name in sorted(names_2026 & names_2030):
        actions_2026 = dict_2026[name]
        actions_2030 = dict_2030[name]

        normalized_2026 = [normalize_action_line(line) for line in actions_2026]
        normalized_2030 = [normalize_action_line(line) for line in actions_2030]

        if normalized_2026 == normalized_2030:
            unchanged.append({"Contingency": name, "Actions": join_actions(actions_2026)})
        else:
            modified.append(
                {
                    "Contingency": name,
                    "Actions_2026": join_actions(actions_2026),
                    "Actions_2030": join_actions(actions_2030),
                }
            )

    for name in sorted(names_2026 - names_2030):
        removed.append({"Contingency": name, "Actions": join_actions(dict_2026[name])})

    for name in sorted(names_2030 - names_2026):
        added.append({"Contingency": name, "Actions": join_actions(dict_2030[name])})

    return {
        "summary": {
            "Unchanged": len(unchanged),
            "Modified": len(modified),
            "Added": len(added),
            "Removed": len(removed),
        },
        "unchanged": unchanged,
        "modified": modified,
        "removed": removed,
        "added": added,
    }


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
STATUS_FILLS = {
    "Added": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "Removed": PatternFill(fill_type="solid", fgColor="FDE9E7"),
    "Modified": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Unchanged": PatternFill(fill_type="solid", fgColor="F2F2F2"),
}
THIN_BOTTOM_BORDER = Border(bottom=Side(style="thin", color="808080"))


def style_header_row(ws, row_number):
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM_BORDER


def apply_fixed_column_widths(ws, widths):
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def write_table(ws, start_row, headers, rows, wrap_columns=None):
    wrap_columns = set(wrap_columns or [])

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
    style_header_row(ws, start_row)

    for row_offset, row_data in enumerate(rows, start=1):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value=row_data.get(header, ""))
            cell.alignment = Alignment(
                wrap_text=header in wrap_columns,
                vertical="top",
                horizontal="left",
            )


def build_side_by_side_rows(results, status):
    if status == "Added":
        source_rows = results["added"]
        rows = [
            {
                "Contingency": row["Contingency"],
                "Actions_2026": "",
                "Actions_2030": row["Actions"],
                "Status": "Added",
            }
            for row in source_rows
        ]
    elif status == "Removed":
        source_rows = results["removed"]
        rows = [
            {
                "Contingency": row["Contingency"],
                "Actions_2026": row["Actions"],
                "Actions_2030": "",
                "Status": "Removed",
            }
            for row in source_rows
        ]
    elif status == "Unchanged":
        source_rows = results["unchanged"]
        rows = [
            {
                "Contingency": row["Contingency"],
                "Actions_2026": row["Actions"],
                "Actions_2030": row["Actions"],
                "Status": "Unchanged",
            }
            for row in source_rows
        ]
    else:
        source_rows = results["modified"]
        rows = [
            {
                "Contingency": row["Contingency"],
                "Actions_2026": row["Actions_2026"],
                "Actions_2030": row["Actions_2030"],
                "Status": "Modified",
            }
            for row in source_rows
        ]

    return sorted(rows, key=lambda row: row["Contingency"])


def apply_status_row_fill(ws, start_row, status_column_index, row_count):
    for row_idx in range(start_row + 1, start_row + 1 + row_count):
        status = ws.cell(row=row_idx, column=status_column_index).value
        fill = STATUS_FILLS.get(status)
        if fill is None:
            continue
        for col_idx in range(1, status_column_index + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def write_summary_sheet(ws, prefix, results, file_paths):
    ws.title = f"{prefix}_Summary"
    metric_rows = [{"Metric": metric, "Count": count} for metric, count in results["summary"].items()]
    write_table(ws, start_row=1, headers=["Metric", "Count"], rows=metric_rows)
    ws.freeze_panes = "A2"

    how_to_read_row = len(metric_rows) + 3
    ws.cell(
        row=how_to_read_row,
        column=1,
        value="HOW TO READ: Actions_2026 is left, Actions_2030 is right.",
    )
    ws.cell(row=how_to_read_row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=how_to_read_row, column=1).alignment = Alignment(horizontal="left", vertical="top")

    note_rows = [
        {
            "Status": "Added",
            "Meaning": "Exists only in 2030. There is no matching contingency in 2026.",
        },
        {
            "Status": "Removed",
            "Meaning": "Exists only in 2026. It is not present in 2030.",
        },
        {
            "Status": "Modified",
            "Meaning": "Exists in both years, but the action lines changed.",
        },
        {
            "Status": "Unchanged",
            "Meaning": "Exists in both years, and the action lines match.",
        },
    ]
    write_table(
        ws,
        start_row=how_to_read_row + 2,
        headers=["Status", "Meaning"],
        rows=note_rows,
        wrap_columns={"Meaning"},
    )
    apply_fixed_column_widths(ws, {"A": 18, "B": 100})

    input_rows = [
        {"Label": "Generated", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Label": f"{prefix} 2026", "Value": file_paths[f"{prefix}_2026"]},
        {"Label": f"{prefix} 2030", "Value": file_paths[f"{prefix}_2030"]},
    ]
    inputs_start_row = how_to_read_row + 2 + len(note_rows) + 2
    write_table(ws, start_row=inputs_start_row, headers=["Label", "Value"], rows=input_rows, wrap_columns={"Value"})
    apply_fixed_column_widths(ws, {"A": 18, "B": 100})


def write_start_here_sheet(ws, file_paths):
    ws.title = "Start_Here"
    ws.freeze_panes = "A2"
    ws["A1"] = "Contingency Compare Workbook"
    ws["A1"].font = Font(bold=True, size=14)

    rows = [
        ("What was compared", ""),
        ("P1 2026", file_paths["P1_2026"]),
        ("P1 2030", file_paths["P1_2030"]),
        ("P2 2026", file_paths["P2_2026"]),
        ("P2 2030", file_paths["P2_2030"]),
        ("", ""),
        ("Status meaning", ""),
        ("Added", "Only in 2030. Left side blank, right side filled."),
        ("Removed", "Only in 2026. Left side filled, right side blank."),
        ("Modified", "Exists in both years, but actions changed."),
        ("Unchanged", "Exists in both years, and actions match."),
        ("", ""),
        ("Example", ""),
        ("Reading a row", "Compare Actions_2026 on the left to Actions_2030 on the right, then confirm Status."),
    ]

    for row_idx, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="top")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if value == "" and label:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

    apply_fixed_column_widths(ws, {"A": 24, "B": 100})


def write_detail_sheet(ws, title, rows):
    ws.title = title
    headers = ["Contingency", "Actions_2026", "Actions_2030", "Status"]
    write_table(
        ws,
        start_row=1,
        headers=headers,
        rows=rows,
        wrap_columns={"Actions_2026", "Actions_2030"},
    )
    apply_fixed_column_widths(ws, {"A": 45, "B": 70, "C": 70, "D": 12})
    apply_status_row_fill(ws, start_row=1, status_column_index=4, row_count=len(rows))
    ws.freeze_panes = "A2"


def write_inputs_sheet(ws, file_paths):
    ws.title = "Inputs"
    rows = []
    for slot in SLOTS:
        path = file_paths.get(slot, "")
        rows.append(
            {
                "Slot": slot,
                "Filename": Path(path).name if path else "",
                "Full_Path": path,
            }
        )

    write_table(
        ws,
        start_row=1,
        headers=["Slot", "Filename", "Full_Path"],
        rows=rows,
        wrap_columns={"Full_Path"},
    )
    apply_fixed_column_widths(ws, {"A": 18, "B": 40, "C": 100})
    ws.freeze_panes = "A2"


def write_excel_report(out_path, results_p1, results_p2, file_paths):
    workbook = Workbook()
    write_start_here_sheet(workbook.active, file_paths)

    for prefix, results in (("P1", results_p1), ("P2", results_p2)):
        write_summary_sheet(workbook.create_sheet(), prefix, results, file_paths)
        write_detail_sheet(workbook.create_sheet(), f"{prefix}_Added", build_side_by_side_rows(results, "Added"))
        write_detail_sheet(workbook.create_sheet(), f"{prefix}_Removed", build_side_by_side_rows(results, "Removed"))
        write_detail_sheet(
            workbook.create_sheet(),
            f"{prefix}_Modified",
            build_side_by_side_rows(results, "Modified"),
        )
        write_detail_sheet(
            workbook.create_sheet(),
            f"{prefix}_Unchanged",
            build_side_by_side_rows(results, "Unchanged"),
        )

    write_inputs_sheet(workbook.create_sheet(), file_paths)
    workbook.save(out_path)


class ContingencyCompareApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Contingency Compare")
        self.root.geometry("860x520")
        self.root.minsize(760, 460)

        self.selected_folder = ""
        self.file_paths = {}
        self.path_vars = {slot: tk.StringVar(value="Not found") for slot in SLOTS}

        self._build_ui()

    def _build_ui(self):
        container = ttk.Frame(self.root, padding=16)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)
        container.rowconfigure(6, weight=1)

        title = ttk.Label(container, text="Contingency Compare", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))

        select_button = tk.Button(
            container,
            text="Select Folder",
            font=("Segoe UI", 14, "bold"),
            padx=16,
            pady=12,
            command=self.select_folder,
        )
        select_button.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 16))

        for row_idx, slot in enumerate(SLOTS, start=2):
            ttk.Label(container, text=slot.replace("_", " "), width=12).grid(row=row_idx, column=0, sticky="w", pady=4)
            entry = ttk.Entry(container, textvariable=self.path_vars[slot], state="readonly")
            entry.grid(row=row_idx, column=1, sticky="ew", pady=4)

        self.status_label = ttk.Label(container, text="Select a folder to begin.", foreground="#333333")
        self.status_label.grid(row=6, column=0, columnspan=2, sticky="w", pady=(12, 4))

        self.status_text = tk.Text(container, height=10, wrap="word", state="disabled")
        self.status_text.grid(row=7, column=0, columnspan=2, sticky="nsew", pady=(0, 12))

        self.run_button = ttk.Button(container, text="Run Compare", command=self.run_compare, state="disabled")
        self.run_button.grid(row=8, column=0, columnspan=2, sticky="ew")

    def set_status(self, message, append=False, error=False):
        self.status_label.config(
            text=message.splitlines()[0] if message else "",
            foreground="#b00020" if error else "#333333",
        )
        self.status_text.config(state="normal")
        if not append:
            self.status_text.delete("1.0", tk.END)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state="disabled")
        self.root.update_idletasks()

    def select_folder(self):
        folder = filedialog.askdirectory()
        if not folder:
            return

        self.selected_folder = folder
        detection = find_con_files(folder)
        self.file_paths = detection["selected"]

        for slot in SLOTS:
            path = self.file_paths.get(slot, "")
            self.path_vars[slot].set(path if path else "Not found")

        messages = [f"Selected folder: {folder}"]
        if detection["warnings"]:
            messages.extend(f"Warning: {warning}" for warning in detection["warnings"])

        if detection["missing"]:
            missing_text = ", ".join(detection["missing"])
            messages.append(f"Missing: {missing_text}")
            self.run_button.config(state="disabled")
            self.set_status("\n".join(messages), error=True)
        else:
            messages.append("All required files found.")
            self.run_button.config(state="normal")
            self.set_status("\n".join(messages))

    def run_compare(self):
        if not self.selected_folder or len(self.file_paths) != len(SLOTS):
            self.set_status("Missing required files. Select a valid folder first.", error=True)
            self.run_button.config(state="disabled")
            return

        try:
            self.set_status("Parsing...")
            p1_2026 = parse_con_file(self.file_paths["P1_2026"])
            p1_2030 = parse_con_file(self.file_paths["P1_2030"])
            p2_2026 = parse_con_file(self.file_paths["P2_2026"])
            p2_2030 = parse_con_file(self.file_paths["P2_2030"])

            self.set_status("Comparing...", append=True)
            results_p1 = compare_sets(p1_2026, p1_2030)
            results_p2 = compare_sets(p2_2026, p2_2030)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = os.path.join(self.selected_folder, f"Contingency_Compare_{timestamp}.xlsx")

            self.set_status("Writing report...", append=True)
            write_excel_report(output_path, results_p1, results_p2, self.file_paths)

            self.set_status(f"Done: {output_path}", append=True)
            self._show_completion(output_path)
        except Exception as exc:
            self.set_status(f"Error: {exc}", append=True, error=True)
            messagebox.showerror("Contingency Compare", str(exc))

    def _show_completion(self, output_path):
        messagebox.showinfo("Contingency Compare", f"Report written to:\n{output_path}")
        if messagebox.askyesno("Open Folder", "Open the output folder in File Explorer?"):
            self.open_output_location(output_path)

    @staticmethod
    def open_output_location(output_path):
        folder = os.path.dirname(output_path)
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception:
            pass


def main():
    root = tk.Tk()
    ttk.Style().theme_use("clam")
    app = ContingencyCompareApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
