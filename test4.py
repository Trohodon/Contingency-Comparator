import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COL_ALIASES = {
    "from": ["From", "FROM", "From kV", "From_kV", "From KV", "Prim_kV", "Prim KV"],
    "to": ["To", "TO", "To kV", "To_kV", "To KV", "Sec_kV", "Sec KV"],
    "core": [
        "Core Loss", "Core Losses", "Core_Loss", "Core_Losses",
        "Typical Core Losses (Watts)", "Typical Core Losses", "Core Losses (Watts)",
        "Core Losses W", "CoreLoss_W", "Core Losses (W)"
    ],
}


DOM_HEADER_FILL = PatternFill("solid", fgColor="002F6C")   # dark blue
DOM_HEADER_FONT = Font(color="FFFFFF", bold=True)
DOM_BAND_FILL = PatternFill("solid", fgColor="F2F2F2")     # light gray
DOM_TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")    # pale blue-gray
DOM_TEXT_FONT = Font(color="000000", bold=False)

THIN = Side(style="thin", color="9E9E9E")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _norm(s: str) -> str:
    return str(s).strip().lower()


def _pick_col(df: pd.DataFrame, key: str) -> str:
    cols = list(df.columns)
    cols_norm = {_norm(c): c for c in cols}
    for cand in COL_ALIASES[key]:
        ckey = _norm(cand)
        if ckey in cols_norm:
            return cols_norm[ckey]
    raise KeyError(f"Missing required column for '{key}'. Found columns: {cols}")


def _to_num(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace(",", "", regex=False).str.strip()
    s = s.replace({"": np.nan, "nan": np.nan, "None": np.nan})
    return pd.to_numeric(s, errors="coerce")


def read_sheet_with_auto_header(xlsx_path: str, sheet_name: str, scan_rows: int = 250) -> pd.DataFrame:
    preview = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, nrows=scan_rows, dtype=str)
    preview = preview.fillna("").astype(str)
    preview = preview.apply(lambda col: col.str.strip())

    need_from = {_norm(x) for x in COL_ALIASES["from"]}
    need_to = {_norm(x) for x in COL_ALIASES["to"]}
    need_core = {_norm(x) for x in COL_ALIASES["core"]}

    header_row = None
    for r in range(len(preview)):
        row_vals = {_norm(v) for v in preview.iloc[r].tolist() if v}
        if (row_vals & need_from) and (row_vals & need_to) and (row_vals & need_core):
            header_row = r
            break

    if header_row is None:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
        return df

    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row)
    df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")]
    return df


def summarize_core_losses(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    c_from = _pick_col(df, "from")
    c_to = _pick_col(df, "to")
    c_core = _pick_col(df, "core")

    from_kv = _to_num(df[c_from])
    to_kv = _to_num(df[c_to])
    core_w = _to_num(df[c_core]).fillna(0.0)

    prim = np.fmax(from_kv, to_kv)
    sec = np.fmin(from_kv, to_kv)

    out = pd.DataFrame({"Prim_kV": prim, "Sec_kV": sec, "CoreLoss_W": core_w})
    out = out.dropna(subset=["Prim_kV", "Sec_kV"])

    summary = (
        out.groupby(["Prim_kV", "Sec_kV"], dropna=False)["CoreLoss_W"]
           .sum()
           .reset_index()
           .rename(columns={"CoreLoss_W": "SumOfCoreLosses_W"})
    )

    summary["Prim_kV"] = summary["Prim_kV"].astype(float)
    summary["Sec_kV"] = summary["Sec_kV"].astype(float)
    summary["SumOfCoreLosses_W"] = summary["SumOfCoreLosses_W"].round(0).astype(np.int64)

    summary = summary.sort_values(["Prim_kV", "Sec_kV"], ascending=[False, False]).reset_index(drop=True)

    total_row = pd.DataFrame({
        "Prim_kV": [np.nan],
        "Sec_kV": [np.nan],
        "SumOfCoreLosses_W": [int(summary["SumOfCoreLosses_W"].sum())]
    })
    summary = pd.concat([summary, total_row], ignore_index=True)
    return summary


def _autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 28)


def format_summary_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_row = 1
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = DOM_HEADER_FILL
        cell.font = DOM_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    last_row = ws.max_row
    last_col = ws.max_column

    for r in range(2, last_row + 1):
        is_total = (r == last_row)
        band = (r % 2 == 0)

        row_fill = DOM_TOTAL_FILL if is_total else (DOM_BAND_FILL if band else None)

        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if row_fill is not None:
                cell.fill = row_fill
            cell.font = DOM_TEXT_FONT
            cell.border = BORDER_THIN

            if c in (1, 2):
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 3:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.cell(row=last_row, column=1).value = "Total"
    ws.cell(row=last_row, column=1).font = Font(bold=True, color="000000")
    ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=last_row, column=2).value = ""
    ws.cell(row=last_row, column=2).border = BORDER_THIN

    ws.cell(row=last_row, column=3).font = Font(bold=True, color="000000")

    _autosize_columns(ws)


def run_process(input_path: str, sheet_name: str, output_path: str) -> None:
    df = read_sheet_with_auto_header(input_path, sheet_name)
    summary = summarize_core_losses(df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="Summary", index=False)
        ws = w.book["Summary"]
        format_summary_sheet(ws)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Core Loss Summarizer")
        self.geometry("860x300")

        self.in_path = tk.StringVar()
        self.out_path = tk.StringVar()
        self.sheet_name = tk.StringVar()

        pad = {"padx": 10, "pady": 6}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ttk.Label(frm, text="Input .xlsx").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.in_path, width=78).grid(row=0, column=1, sticky="we", **pad)
        ttk.Button(frm, text="Browse", command=self.pick_input).grid(row=0, column=2, **pad)

        ttk.Label(frm, text="Sheet").grid(row=1, column=0, sticky="w", **pad)
        self.sheet_combo = ttk.Combobox(frm, textvariable=self.sheet_name, width=48, state="readonly", values=[])
        self.sheet_combo.grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(frm, text="Output .xlsx").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.out_path, width=78).grid(row=2, column=1, sticky="we", **pad)
        ttk.Button(frm, text="Browse", command=self.pick_output).grid(row=2, column=2, **pad)

        ttk.Separator(frm).grid(row=3, column=0, columnspan=3, sticky="we", pady=10)
        ttk.Button(frm, text="Run", command=self.run).grid(row=4, column=1, sticky="e", padx=10, pady=10)

        frm.columnconfigure(1, weight=1)

    def pick_input(self):
        path = filedialog.askopenfilename(
            title="Select input Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return

        self.in_path.set(path)
        try:
            xl = pd.ExcelFile(path)
            sheets = xl.sheet_names
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read workbook sheets:\n{e}")
            return

        self.sheet_combo["values"] = sheets
        if sheets:
            self.sheet_name.set(sheets[0])

        if not self.out_path.get():
            self.out_path.set(path.replace(".xlsx", "_coreloss_summary.xlsx"))

    def pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Save output Excel file as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.out_path.set(path)

    def run(self):
        in_path = self.in_path.get().strip()
        out_path = self.out_path.get().strip()
        sheet = self.sheet_name.get().strip()

        if not in_path:
            messagebox.showerror("Missing input", "Pick an input .xlsx file.")
            return
        if not sheet:
            messagebox.showerror("Missing sheet", "Pick a sheet from the dropdown.")
            return
        if not out_path:
            messagebox.showerror("Missing output", "Pick an output .xlsx file.")
            return

        try:
            run_process(in_path, sheet, out_path)
        except Exception as e:
            messagebox.showerror("Failed", str(e))
            return

        messagebox.showinfo("Done", f"Saved:\n{out_path}")


if __name__ == "__main__":
    App().mainloop()